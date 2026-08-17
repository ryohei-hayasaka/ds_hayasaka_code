"""Non-interactive Tk smoke test used before the portable GUI walkthrough."""

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from tga_analyzer.excel_export import export_excel
from tga_analyzer.gui import TgaAnalyzerApp
from tga_analyzer.model import DSC, GPC, IR, PARTICLE_SIZE, TGA, UV_VIS
from tga_analyzer.parser import (
    load_dsc_csv,
    load_gpc_csv,
    load_ir_csv,
    load_particle_size_csv,
    load_uvvis_csv,
)
from tga_analyzer.particle_size_processing import ParticleSizeSeriesSettings
from tga_analyzer.processing import USE_NONE


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    app = TgaAnalyzerApp()
    try:
        app.withdraw()
        app.mode_var.set(IR)
        app._on_mode_changed()
        ir_folder = root / "demo_data" / "IR" / "raw_data"
        ir_sample = load_ir_csv(ir_folder / "IR_demo_01.csv")
        ir_blank = load_ir_csv(ir_folder / "IR_blank.csv")
        app.states[IR].add_curve(ir_sample)
        app.states[IR].add_curve(ir_blank)
        app.common_processing[IR].blank_key = ir_blank.key
        app.common_processing[IR].normalization_wavenumber = 1600.0
        app._reprocess_mode(IR)
        ir_result = app.processed_curves[IR][ir_sample.key]
        assert ir_result.status == "Normalized"
        assert ir_result.point_count == ir_sample.point_count
        assert app.loaded_tree.set("curve_0", "blank") == ir_blank.display_name
        assert app.loaded_tree.set("curve_0", "normalization") == "1600 cm⁻¹"

        graph = app._open_graph_window()
        graph.withdraw()
        app.update_idletasks()
        assert len(graph.plot_canvas._curves) == 2
        assert graph.plot_canvas._measurement_type == IR
        assert graph.loaded_tree.set("curve_0", "blank") == ir_blank.display_name
        assert graph.loaded_tree.set("curve_0", "normalization") == "1600 cm⁻¹"

        app.mode_var.set(DSC)
        app._on_mode_changed()
        dsc_folder = root / "demo_data" / "DSC" / "raw_data"
        dsc_sample = load_dsc_csv(dsc_folder / "DSC_demo_01.csv")
        dsc_blank = load_dsc_csv(dsc_folder / "DSC_blank.csv")
        app.states[DSC].add_curve(dsc_sample)
        app.states[DSC].add_curve(dsc_blank)
        app.common_processing[DSC].blank_key = dsc_blank.key
        app._reprocess_mode(DSC)
        dsc_result = app.processed_curves[DSC][dsc_sample.key]
        assert dsc_result.status == "Blank corrected"
        assert app._dsc_analysis_curve(dsc_sample.key).heat_flow_mw == dsc_result.display_y
        assert app.loaded_tree.set("curve_0", "blank") == dsc_blank.display_name
        assert "normalization" not in app.loaded_tree.cget("displaycolumns")

        app.mode_var.set(TGA)
        app._on_mode_changed()
        app.update_idletasks()
        assert app.state_model.measurement_type == TGA
        assert "blank" not in app.loaded_tree.cget("displaycolumns")
        assert "normalization" not in app.loaded_tree.cget("displaycolumns")

        app.mode_var.set(UV_VIS)
        app._on_mode_changed()
        uv_folder = root / "demo_data" / "UV-Vis" / "raw_data"
        uv_first = load_uvvis_csv(uv_folder / "UVVis_demo_01.csv")
        uv_second = load_uvvis_csv(uv_folder / "UVVis_demo_02.csv")
        app.states[UV_VIS].add_curve(uv_first)
        app.states[UV_VIS].add_curve(uv_second)
        app.states[UV_VIS].set_color(uv_first.key, "#123456")
        app.states[UV_VIS].set_legend_name(uv_first.key, "UV smoke")
        app._refresh_loaded_curves()
        app.update_idletasks()
        assert app.state_model.measurement_type == UV_VIS
        assert len(graph.plot_canvas._curves) == 2
        assert graph.plot_canvas._curves[0].x_values == uv_first.wavelengths_nm
        assert graph.plot_canvas._curves[0].y_values == uv_first.uvvis_absorbance
        assert graph.plot_canvas._curves[0].legend_name == "UV smoke"
        assert graph.plot_canvas._curves[0].color == "#123456"
        assert not graph.plot_canvas._curves[0].reverse_x
        assert "blank" not in app.loaded_tree.cget("displaycolumns")
        assert "normalization" not in app.loaded_tree.cget("displaycolumns")
        assert "blank" not in graph.loaded_tree.cget("displaycolumns")
        assert "normalization" not in graph.loaded_tree.cget("displaycolumns")

        app.mode_var.set(GPC)
        app._on_mode_changed()
        gpc_folder = root / "demo_data" / "GPC" / "raw_data"
        gpc_files = (
            "GPC_RI_demo_01.csv",
            "GPC_RI_demo_07.csv",
            "GPC_RI_demo_10.csv",
        )
        gpc_curves = tuple(load_gpc_csv(gpc_folder / name) for name in gpc_files)
        for curve in gpc_curves:
            app.states[GPC].add_curve(curve)
        app.states[GPC].set_legend_name(gpc_curves[1].key, "GPC double peak")
        app._refresh_loaded_curves()
        app.update_idletasks()
        assert app.state_model.measurement_type == GPC
        assert len(graph.plot_canvas._curves) == 3
        assert graph.plot_canvas._curves[1].x_values == gpc_curves[1].retention_times_min
        assert graph.plot_canvas._curves[1].y_values == gpc_curves[1].ri_signal_mv
        assert graph.plot_canvas._curves[1].legend_name == "GPC double peak"
        assert not graph.plot_canvas._curves[1].reverse_x

        app.mode_var.set(PARTICLE_SIZE)
        app._on_mode_changed()
        particle_folder = root / "demo_data" / "ParticleSize" / "raw_data"
        particle_first = load_particle_size_csv(
            particle_folder / "ParticleSize_demo_01.csv"
        )
        particle_second = load_particle_size_csv(
            particle_folder / "ParticleSize_demo_07.csv"
        )
        app.states[PARTICLE_SIZE].add_curve(particle_first)
        app.states[PARTICLE_SIZE].add_curve(particle_second)
        app.particle_common_processing.normalization_diameter_um = 1.0
        app.particle_series_processing[particle_second.key] = (
            ParticleSizeSeriesSettings(normalization_mode=USE_NONE)
        )
        app._reprocess_particle_size()
        app.update_idletasks()
        particle_result = app.particle_processed_curves[particle_first.key]
        reference_index = particle_first.particle_diameter_um.index(1.0)
        assert particle_result.display_y[reference_index] == 1.0
        assert app.loaded_tree.set("curve_0", "particle_normalization") == "1 µm"
        assert app.loaded_tree.set("curve_1", "particle_normalization") == "なし"
        assert graph.loaded_tree.set("curve_0", "particle_normalization") == "1 µm"
        assert len(graph.plot_canvas._curves) == 2
        assert graph.plot_canvas._curves[0].logarithmic_x
        assert not graph.plot_canvas._curves[0].reverse_x
        assert graph.plot_canvas._y_axis_title() == "Volume / Normalized volume"
        assert app.states[PARTICLE_SIZE].axis_range.x_min == 0.1
        assert app.states[PARTICLE_SIZE].axis_range.x_max == 1000.0

        with tempfile.TemporaryDirectory() as temp_dir:
            for mode, expected_rows, x_header, y_header in (
                (UV_VIS, 602, "Wavelength_nm", "Absorbance"),
                (GPC, 1502, "RetentionTime_min", "RI_mV"),
                (PARTICLE_SIZE, 162, "ParticleDiameter_um", "NormalizedVolume"),
            ):
                state = app.states[mode]
                output = export_excel(
                    app._display_series(mode),
                    state,
                    Path(temp_dir) / f"{mode}_smoke.xlsx",
                )
                sheet = load_workbook(output)["Data"]
                assert sheet.max_row == expected_rows
                assert x_header in sheet["A1"].value
                assert y_header in sheet["B1"].value
                assert len(sheet._charts) == 1
                assert sheet._charts[0].anchor._from.col == 3
                assert sheet._charts[0].anchor._from.row == 4
                if mode == PARTICLE_SIZE:
                    assert sheet._charts[0].x_axis.scaling.logBase == 10.0

        app.mode_var.set(IR)
        app._on_mode_changed()
        app.update_idletasks()
        assert app.states[IR].curves[ir_sample.key].color != "#123456"
        assert app.state_model.measurement_type == IR
        print(
            "GUI source smoke passed: six isolated modes, raw UV-Vis/GPC and "
            "log particle-size plots, processing, graph window, native Excel"
        )
    finally:
        app._on_close()


if __name__ == "__main__":
    main()
