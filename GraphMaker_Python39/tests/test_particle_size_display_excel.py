import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from tga_analyzer.display_series import display_axis_titles, to_display_series
from tga_analyzer.excel_export import export_excel
from tga_analyzer.model import PARTICLE_SIZE, AxisRange, PlotState
from tga_analyzer.parser import load_particle_size_csv
from tga_analyzer.particle_size_processing import (
    ParticleSizeCommonSettings,
    ParticleSizeSeriesSettings,
    process_particle_size_curve,
)
from tga_analyzer.processing import USE_INDIVIDUAL, USE_NONE


ROOT = Path(__file__).resolve().parents[1]
DEMO = ROOT / "demo_data" / "ParticleSize" / "raw_data" / "ParticleSize_demo_07.csv"


def axis_title(axis) -> str:
    return "".join(
        run.t
        for paragraph in axis.title.tx.rich.paragraphs
        for run in paragraph.r
    )


class ParticleSizeDisplayExcelTests(unittest.TestCase):
    def test_display_projection_uses_log_x_and_dynamic_y_titles(self):
        curve = load_particle_size_csv(DEMO)
        raw = process_particle_size_curve(
            curve,
            ParticleSizeCommonSettings(),
            ParticleSizeSeriesSettings(normalization_mode=USE_NONE),
        )
        normalized = process_particle_size_curve(
            curve,
            ParticleSizeCommonSettings(),
            ParticleSizeSeriesSettings(
                normalization_mode=USE_INDIVIDUAL,
                normalization_diameter_um=10.0,
            ),
        )
        raw_display = to_display_series(raw)
        normalized_display = to_display_series(normalized)
        self.assertTrue(raw_display.logarithmic_x)
        self.assertFalse(raw_display.reverse_x)
        self.assertEqual(raw_display.y_data_header, "VolumeFrequency_percent")
        self.assertEqual(normalized_display.y_data_header, "NormalizedVolume")
        self.assertEqual(normalized_display.y_axis_title, "Normalized volume")
        self.assertEqual(
            display_axis_titles((raw_display, normalized_display), PARTICLE_SIZE)[1],
            "Volume / Normalized volume",
        )

    def test_excel_exports_current_normalized_data_with_native_log_chart(self):
        curve = load_particle_size_csv(DEMO)
        curve.legend_name = "Particle edited"
        curve.color = "#123456"
        processed = process_particle_size_curve(
            curve,
            ParticleSizeCommonSettings(),
            ParticleSizeSeriesSettings(
                normalization_mode=USE_INDIVIDUAL,
                normalization_diameter_um=10.0,
            ),
        )
        state = PlotState(measurement_type=PARTICLE_SIZE)
        state.add_curve(curve)
        state.set_color(curve.key, "#123456")
        state.set_legend_name(curve.key, "Particle edited")
        state.set_manual_range(AxisRange(0.1, 1000.0, 0.0, 5.0))
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel(
                (processed,), state, Path(temp_dir) / "particle_size.xlsx"
            )
            sheet = load_workbook(output)["Data"]
            with ZipFile(output) as archive:
                chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")
        self.assertEqual(sheet.max_row, 162)
        self.assertIn("Particle edited | ParticleDiameter_um", sheet["A1"].value)
        self.assertIn("Particle edited | NormalizedVolume", sheet["B1"].value)
        self.assertIn("Reference=10 um", sheet["B1"].value)
        self.assertEqual(sheet["B2"].value, processed.display_y[0])
        chart = sheet._charts[0]
        self.assertEqual(chart.anchor._from.col, 3)
        self.assertEqual(chart.anchor._from.row, 4)
        self.assertEqual(chart.x_axis.scaling.logBase, 10.0)
        self.assertEqual(chart.x_axis.scaling.min, 0.1)
        self.assertEqual(chart.x_axis.scaling.max, 1000.0)
        self.assertEqual(axis_title(chart.x_axis), "Particle diameter (µm)")
        self.assertEqual(axis_title(chart.y_axis), "Normalized volume")
        self.assertIn('<symbol val="none"', chart_xml)
        self.assertEqual(chart.series[0].tx.v, "Particle edited")

    def test_failed_normalization_exports_raw_values_and_reason(self):
        curve = load_particle_size_csv(DEMO)
        processed = process_particle_size_curve(
            curve,
            ParticleSizeCommonSettings(),
            ParticleSizeSeriesSettings(
                normalization_mode=USE_INDIVIDUAL,
                normalization_diameter_um=0.01,
            ),
        )
        state = PlotState(measurement_type=PARTICLE_SIZE)
        state.add_curve(curve)
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((processed,), state, Path(temp_dir) / "failed.xlsx")
            sheet = load_workbook(output)["Data"]
        self.assertIn("VolumeFrequency_percent", sheet["B1"].value)
        self.assertIn("not applied", sheet["B1"].value)
        self.assertIn("範囲外", sheet["B1"].value)
        self.assertEqual(sheet["B2"].value, curve.volume_frequency_percent[0])


if __name__ == "__main__":
    unittest.main()
