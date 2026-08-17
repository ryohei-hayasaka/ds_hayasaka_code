import tempfile
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.units import EMU_to_cm

from tga_analyzer.excel_export import export_excel
from tga_analyzer.model import DSC, IR, AxisRange, CurveData, PlotState
from tga_analyzer.processing import (
    CommonProcessingSettings,
    SeriesProcessingSettings,
    process_dsc_curve,
    process_ir_curve,
)


def make_curve(name: str, color: str, offset: float = 0.0) -> CurveData:
    return CurveData(
        path=Path(name + ".csv"),
        display_name=name,
        temperatures=(25.0, 100.0, 800.0),
        mass_mg=(10.0 + offset, 9.0 + offset, 2.0 + offset),
        weight_percent=(100.0, 90.0, 20.0),
        color=color,
    )


class ExcelExportTests(unittest.TestCase):
    def test_exports_editable_chart_with_series_colors_and_axis_range(self):
        first = make_curve("Sample A", "#112233")
        second = make_curve("Sample B", "#AABBCC", 1.0)
        state = PlotState(
            curves={first.key: first, second.key: second},
            axis_range=AxisRange(100, 700, 10, 100),
            auto_axes=False,
        )
        state.set_legend_name(first.key, "Edited A")
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((first, second), state, Path(temp_dir) / "result.xlsx")
            workbook = load_workbook(output)
            with ZipFile(output) as archive:
                chart_xml = ET.fromstring(archive.read("xl/charts/chart1.xml"))

        self.assertEqual(workbook.properties.creator, "GraphMaker Python 3.9 Compatible")
        self.assertEqual(workbook.sheetnames, ["Data"])
        data = workbook["Data"]
        self.assertEqual(data["A1"].value, "Edited A | Temperature_C")
        self.assertFalse(data["A1"].font.bold)
        self.assertIsNone(data["A1"].fill.fill_type)
        self.assertIsNone(data["A1"].border.left.style)
        self.assertIsNone(data["A1"].alignment.wrap_text)
        self.assertIsNone(data["A1"].comment)
        self.assertIsNone(data.row_dimensions[1].height)
        self.assertEqual(data.column_dimensions["A"].width, 7.0)
        self.assertEqual(data["B2"].value, 100.0)
        self.assertEqual(data["D1"].value, "Sample B | Temperature_C")
        self.assertEqual(data.freeze_panes, "A2")
        self.assertEqual(len(data._charts), 1)
        chart = data._charts[0]
        self.assertEqual(chart.anchor._from.col, 3)
        self.assertEqual(chart.anchor._from.row, 4)
        self.assertAlmostEqual(EMU_to_cm(chart.anchor.ext.cx), 18.0, places=2)
        self.assertAlmostEqual(EMU_to_cm(chart.anchor.ext.cy), 12.0, places=2)
        self.assertEqual(len(chart.series), 2)
        self.assertEqual(chart.x_axis.scaling.min, 100.0)
        self.assertEqual(chart.x_axis.scaling.max, 700.0)
        self.assertEqual(chart.y_axis.scaling.min, 10.0)
        self.assertEqual(chart.y_axis.scaling.max, 100.0)
        self.assertEqual(chart.x_axis.tickLblPos, "low")
        self.assertEqual(chart.y_axis.tickLblPos, "nextTo")
        self.assertEqual(chart.x_axis.axPos, "b")
        self.assertEqual(chart.y_axis.axPos, "l")
        self.assertEqual(chart.x_axis.crosses, "min")
        self.assertEqual(chart.y_axis.crosses, "min")
        self.assertIsNone(chart.x_axis.majorGridlines)
        self.assertIsNone(chart.y_axis.majorGridlines)
        self.assertEqual(chart.x_axis.majorTickMark, "in")
        self.assertEqual(chart.y_axis.majorTickMark, "in")
        self.assertFalse(chart.varyColors)
        self.assertIsNone(chart.title)
        self.assertLess(chart.layout.manualLayout.x, 0.11)
        self.assertLess(chart.layout.manualLayout.y, 0.05)
        self.assertGreater(chart.layout.manualLayout.w, 0.75)
        self.assertGreater(chart.layout.manualLayout.h, 0.82)
        self.assertEqual(
            chart.graphical_properties.solidFill.srgbClr,
            "FFFFFF",
        )
        self.assertFalse(chart.legend.overlay)
        self.assertEqual(chart.series[0].tx.v, "Edited A")
        namespaces = {
            "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        plot_properties = chart_xml.find(".//c:plotArea/c:spPr", namespaces)
        self.assertIsNotNone(plot_properties)
        plot_fill = plot_properties.find("a:solidFill/a:srgbClr", namespaces)
        self.assertIsNotNone(plot_fill)
        self.assertEqual(plot_fill.attrib["val"], "FFFFFF")
        plot_border = plot_properties.find("a:ln/a:solidFill/a:srgbClr", namespaces)
        self.assertIsNotNone(plot_border)
        self.assertEqual(plot_border.attrib["val"], "000000")
        self.assertEqual(plot_properties.find("a:ln", namespaces).attrib["w"], "19050")
        self.assertIsNone(chart_xml.find("./c:chart/c:title", namespaces))
        chart_area_properties = chart_xml.find("./c:spPr", namespaces)
        self.assertIsNotNone(chart_area_properties)
        chart_fill = chart_area_properties.find("a:solidFill/a:srgbClr", namespaces)
        self.assertIsNotNone(chart_fill)
        self.assertEqual(chart_fill.attrib["val"], "FFFFFF")
        axis_lines = chart_xml.findall(".//c:valAx/c:spPr/a:ln", namespaces)
        self.assertEqual(len(axis_lines), 2)
        self.assertEqual({line.attrib.get("w") for line in axis_lines}, {"19050"})
        self.assertEqual(
            {
                element.attrib.get("val")
                for element in chart_xml.findall(
                    ".//c:valAx/c:majorTickMark", namespaces
                )
            },
            {"in"},
        )
        axis_title_sizes = {
            element.attrib.get("sz")
            for element in chart_xml.findall(
                ".//c:valAx/c:title//a:defRPr", namespaces
            )
        }
        tick_label_sizes = {
            element.attrib.get("sz")
            for element in chart_xml.findall(
                ".//c:valAx/c:txPr//a:defRPr", namespaces
            )
        }
        legend_sizes = {
            element.attrib.get("sz")
            for element in chart_xml.findall(
                ".//c:legend/c:txPr//a:defRPr", namespaces
            )
        }
        self.assertEqual(axis_title_sizes, {"1400"})
        self.assertEqual(tick_label_sizes, {"1200"})
        self.assertEqual(legend_sizes, {"1200"})
        axis_title_bold = {
            element.attrib.get("b")
            for element in chart_xml.findall(
                ".//c:valAx/c:title//a:defRPr", namespaces
            )
        }
        tick_label_bold = {
            element.attrib.get("b")
            for element in chart_xml.findall(
                ".//c:valAx/c:txPr//a:defRPr", namespaces
            )
        }
        legend_bold = {
            element.attrib.get("b")
            for element in chart_xml.findall(
                ".//c:legend/c:txPr//a:defRPr", namespaces
            )
        }
        self.assertEqual(axis_title_bold, {"1"})
        self.assertEqual(tick_label_bold, {"0"})
        self.assertEqual(legend_bold, {"0"})
        self.assertIn("'Data'!$A$2:$A$4", chart.series[0].xVal.numRef.f)
        self.assertIn("'Data'!$B$2:$B$4", chart.series[0].yVal.numRef.f)
        self.assertEqual(
            chart.series[0].graphicalProperties.line.solidFill.srgbClr,
            "112233",
        )

    def test_rejects_export_without_curves(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "曲線がありません"):
                export_excel((), PlotState(), Path(temp_dir) / "empty.xlsx")

    def test_exports_dsc_heat_flow_as_editable_native_chart(self):
        curve = CurveData(
            path=Path("DSC Sample.csv"),
            display_name="DSC Sample",
            temperatures=(25.0, 100.0, 200.0),
            mass_mg=(),
            weight_percent=(),
            color="#445566",
            measurement_type=DSC,
            heat_flow_mw=(-0.25, 1.5, 0.1),
            time_min=(0.0, 2.5, 5.0),
            heat_flow_unit="mW",
            source_heat_flow_header="HeatFlow_mW",
        )
        state = PlotState(
            curves={curve.key: curve},
            axis_range=AxisRange(20, 210, -1, 2),
            auto_axes=False,
            measurement_type=DSC,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((curve,), state, Path(temp_dir) / "dsc.xlsx")
            workbook = load_workbook(output)

        data = workbook["Data"]
        self.assertEqual(data["A1"].value, "DSC Sample | Temperature_C")
        self.assertEqual(data["B1"].value, "DSC Sample | HeatFlow_mW")
        self.assertEqual(data["C1"].value, "DSC Sample | Time_min")
        self.assertEqual(data["B2"].value, -0.25)
        self.assertEqual(data["C4"].value, 5.0)
        self.assertEqual(data.column_dimensions["A"].width, 7.0)
        self.assertEqual(len(data._charts), 1)
        chart = data._charts[0]
        self.assertEqual(len(chart.series), 1)
        self.assertEqual(chart.y_axis.scaling.min, -1.0)
        self.assertEqual(chart.y_axis.scaling.max, 2.0)
        self.assertIn("'Data'!$A$2:$A$4", chart.series[0].xVal.numRef.f)
        self.assertIn("'Data'!$B$2:$B$4", chart.series[0].yVal.numRef.f)
        self.assertEqual(
            chart.series[0].graphicalProperties.line.solidFill.srgbClr,
            "445566",
        )

    def test_exports_exact_processed_ir_data_metadata_and_reversed_axis(self):
        sample = CurveData(
            path=Path("IR Sample.csv"),
            display_name="IR Sample",
            temperatures=(),
            mass_mg=(),
            weight_percent=(),
            color="#13579B",
            measurement_type=IR,
            wavenumbers_cm1=(4000.0, 3000.0, 2000.0),
            absorbance=(0.6, 0.5, 0.4),
            legend_name="Edited IR",
        )
        blank = CurveData(
            path=Path("IR blank.csv"),
            display_name="IR blank",
            temperatures=(),
            mass_mg=(),
            weight_percent=(),
            measurement_type=IR,
            wavenumbers_cm1=(4000.0, 3000.0, 2000.0),
            absorbance=(0.1, 0.1, 0.1),
        )
        processed = process_ir_curve(
            sample,
            CommonProcessingSettings(
                blank_key=blank.key, normalization_wavenumber=3000.0
            ),
            SeriesProcessingSettings(),
            {sample.key: sample, blank.key: blank},
        )
        state = PlotState(
            curves={sample.key: sample},
            axis_range=AxisRange(2000, 4000, 0, 2),
            auto_axes=False,
            measurement_type=IR,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((processed,), state, Path(temp_dir) / "ir.xlsx")
            workbook = load_workbook(output)

        data = workbook["Data"]
        self.assertIn("Edited IR | Wavenumber_cm-1", data["A1"].value)
        self.assertIn("Blank=IR blank", data["A1"].value)
        self.assertIn("Norm=3000 cm-1", data["A1"].value)
        self.assertIn("Status=Normalized", data["B1"].value)
        self.assertEqual(data["A2"].value, processed.display_x[0])
        self.assertEqual(data["B3"].value, processed.display_y[1])
        self.assertEqual(data.column_dimensions["A"].width, 7.0)
        chart = data._charts[0]
        self.assertEqual(chart.x_axis.scaling.orientation, "maxMin")
        self.assertEqual(chart.x_axis.crosses, "min")
        self.assertEqual(chart.y_axis.crosses, "max")
        self.assertEqual(chart.series[0].tx.v, "Edited IR")
        self.assertIn("'Data'!$A$2:$A$4", chart.series[0].xVal.numRef.f)
        self.assertIn("'Data'!$B$2:$B$4", chart.series[0].yVal.numRef.f)

    def test_exports_processed_dsc_curve_not_raw_curve(self):
        sample = CurveData(
            path=Path("sample.csv"), display_name="sample", temperatures=(20.0, 30.0, 40.0),
            mass_mg=(), weight_percent=(), measurement_type=DSC,
            heat_flow_mw=(2.0, 4.0, 6.0), heat_flow_unit="mW",
            source_heat_flow_header="HeatFlow_mW",
        )
        blank = CurveData(
            path=Path("blank.csv"), display_name="blank", temperatures=(20.0, 30.0, 40.0),
            mass_mg=(), weight_percent=(), measurement_type=DSC,
            heat_flow_mw=(1.0, 1.0, 1.0), heat_flow_unit="mW",
            source_heat_flow_header="HeatFlow_mW",
        )
        processed = process_dsc_curve(
            sample, CommonProcessingSettings(blank_key=blank.key),
            SeriesProcessingSettings(), {sample.key: sample, blank.key: blank},
        )
        state = PlotState(
            curves={sample.key: sample}, axis_range=AxisRange(20, 40, 0, 6),
            auto_axes=False, measurement_type=DSC,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((processed,), state, Path(temp_dir) / "dsc_processed.xlsx")
            workbook = load_workbook(output)
        data = workbook["Data"]
        self.assertEqual(data["B2"].value, 1.0)
        self.assertIn("Blank=blank", data["B1"].value)
        self.assertIn("Status=Blank corrected", data["B1"].value)

    def test_failure_exports_the_actual_fallback_data_and_reason(self):
        sample = CurveData(
            path=Path("fallback.csv"), display_name="fallback", temperatures=(),
            mass_mg=(), weight_percent=(), measurement_type=IR,
            wavenumbers_cm1=(4000.0, 3000.0), absorbance=(0.25, 0.5),
        )
        processed = process_ir_curve(
            sample, CommonProcessingSettings(blank_key="deleted"),
            SeriesProcessingSettings(), {sample.key: sample},
        )
        state = PlotState(
            curves={sample.key: sample}, axis_range=AxisRange(3000, 4000, 0, 1),
            auto_axes=False, measurement_type=IR,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output = export_excel((processed,), state, Path(temp_dir) / "fallback.xlsx")
            data = load_workbook(output)["Data"]
        self.assertEqual(data["B2"].value, 0.25)
        self.assertIn("Blank=Failed", data["B1"].value)
        self.assertIn("Status=Blank failed / Raw displayed", data["B1"].value)
        self.assertIn("削除済み", data["B1"].value)


if __name__ == "__main__":
    unittest.main()
