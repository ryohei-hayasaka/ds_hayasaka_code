import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tga_analyzer.display_series import to_display_series
from tga_analyzer.excel_export import export_excel
from tga_analyzer.gui import _load_curve_batch, analysis_panel_for_mode
from tga_analyzer.model import (
    GPC,
    IR,
    MEASUREMENT_TYPES,
    PARTICLE_SIZE,
    TGA,
    UV_VIS,
    AxisRange,
    PlotState,
)
from tga_analyzer.parser import load_gpc_csv, load_uvvis_csv


ROOT = Path(__file__).resolve().parents[1]
UV_FOLDER = ROOT / "demo_data" / "UV-Vis" / "raw_data"
GPC_FOLDER = ROOT / "demo_data" / "GPC" / "raw_data"
PARTICLE_FOLDER = ROOT / "demo_data" / "ParticleSize" / "raw_data"


def _axis_title(axis) -> str:
    return "".join(
        run.t
        for paragraph in axis.title.tx.rich.paragraphs
        for run in paragraph.r
    )


class NewModeStateTests(unittest.TestCase):
    def test_mode_order_and_analysis_panels(self):
        self.assertEqual(
            MEASUREMENT_TYPES,
            ("TGA", "DSC", "IR", "UV-Vis", "GPC", "粒度分布"),
        )
        self.assertEqual(analysis_panel_for_mode(UV_VIS), "raw")
        self.assertEqual(analysis_panel_for_mode(GPC), "raw")
        self.assertEqual(analysis_panel_for_mode(PARTICLE_SIZE), "particle_size")
        self.assertNotEqual(analysis_panel_for_mode(UV_VIS), analysis_panel_for_mode(TGA))
        self.assertNotEqual(analysis_panel_for_mode(GPC), analysis_panel_for_mode(IR))

    def test_batch_loader_uses_mode_specific_parser(self):
        uv_curves, uv_errors = _load_curve_batch(
            [UV_FOLDER / "UVVis_demo_01.csv", UV_FOLDER / "UVVis_demo_02.csv"],
            UV_VIS,
        )
        gpc_curves, gpc_errors = _load_curve_batch(
            [GPC_FOLDER / "GPC_RI_demo_01.csv", GPC_FOLDER / "GPC_RI_demo_07.csv"],
            GPC,
        )
        particle_curves, particle_errors = _load_curve_batch(
            [
                PARTICLE_FOLDER / "ParticleSize_demo_01.csv",
                PARTICLE_FOLDER / "ParticleSize_demo_07.csv",
            ],
            PARTICLE_SIZE,
        )
        self.assertFalse(uv_errors)
        self.assertFalse(gpc_errors)
        self.assertFalse(particle_errors)
        self.assertEqual([curve.measurement_type for curve in uv_curves], [UV_VIS, UV_VIS])
        self.assertEqual([curve.measurement_type for curve in gpc_curves], [GPC, GPC])
        self.assertEqual(
            [curve.measurement_type for curve in particle_curves],
            [PARTICLE_SIZE, PARTICLE_SIZE],
        )

    def test_states_keep_color_legend_order_and_axes_separate(self):
        uv = load_uvvis_csv(UV_FOLDER / "UVVis_demo_01.csv")
        gpc = load_gpc_csv(GPC_FOLDER / "GPC_RI_demo_01.csv")
        uv_state = PlotState(measurement_type=UV_VIS)
        gpc_state = PlotState(measurement_type=GPC)
        uv_state.add_curve(uv)
        gpc_state.add_curve(gpc)
        uv_state.set_color(uv.key, "#123456")
        uv_state.set_legend_name(uv.key, "UV edited")
        uv_state.set_manual_range(AxisRange(250, 700, -0.2, 1.5))
        gpc_state.set_color(gpc.key, "#ABCDEF")
        gpc_state.set_legend_name(gpc.key, "GPC edited")
        gpc_state.set_manual_range(AxisRange(2, 28, -1.0, 8.0))

        uv_reloaded = load_uvvis_csv(UV_FOLDER / "UVVis_demo_01.csv")
        gpc_reloaded = load_gpc_csv(GPC_FOLDER / "GPC_RI_demo_01.csv")
        self.assertTrue(uv_state.replace_curve(uv_reloaded))
        self.assertTrue(gpc_state.replace_curve(gpc_reloaded))

        self.assertEqual(uv_state.ordered_curves()[0].color, "#123456")
        self.assertEqual(uv_state.ordered_curves()[0].legend_label, "UV edited")
        self.assertEqual(gpc_state.ordered_curves()[0].color, "#ABCDEF")
        self.assertEqual(gpc_state.ordered_curves()[0].legend_label, "GPC edited")
        self.assertEqual(uv_state.axis_range, AxisRange(250, 700, -0.2, 1.5))
        self.assertEqual(gpc_state.axis_range, AxisRange(2, 28, -1.0, 8.0))

    def test_auto_range_preserves_negative_raw_values(self):
        uv = load_uvvis_csv(UV_FOLDER / "UVVis_demo_01.csv")
        uv.uvvis_absorbance = tuple(
            -0.25 if index == 0 else value
            for index, value in enumerate(uv.uvvis_absorbance)
        )
        state = PlotState(measurement_type=UV_VIS)
        state.add_curve(uv)
        self.assertLess(state.axis_range.y_min, -0.25)


class NewModeExcelTests(unittest.TestCase):
    def _export(self, curve, mode: str, axis: AxisRange, name: str):
        state = PlotState(measurement_type=mode)
        state.add_curve(curve)
        state.set_color(curve.key, "#2468AC")
        state.set_legend_name(curve.key, f"Edited {mode}")
        state.set_manual_range(axis)
        display = to_display_series(state.ordered_curves()[0])
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        output = export_excel((display,), state, Path(temp_dir.name) / name)
        return load_workbook(output)["Data"], display

    def test_uvvis_excel_keeps_all_601_raw_points(self):
        curve = load_uvvis_csv(UV_FOLDER / "UVVis_demo_01.csv")
        sheet, display = self._export(
            curve, UV_VIS, AxisRange(200, 800, -0.1, 2.0), "uvvis.xlsx"
        )
        chart = sheet._charts[0]
        self.assertEqual(sheet["A1"].value, "Edited UV-Vis | Wavelength_nm")
        self.assertEqual(sheet["B1"].value, "Edited UV-Vis | Absorbance")
        self.assertEqual(sheet.max_row, 602)
        self.assertEqual(sheet["A602"].value, display.x_values[-1])
        self.assertEqual(sheet["B602"].value, display.y_values[-1])
        self.assertIn("'Data'!$A$2:$A$602", chart.series[0].xVal.numRef.f)
        self.assertIn("'Data'!$B$2:$B$602", chart.series[0].yVal.numRef.f)
        self.assertEqual(_axis_title(chart.x_axis), "Wavelength (nm)")
        self.assertEqual(_axis_title(chart.y_axis), "Absorbance")
        self.assertEqual(chart.x_axis.scaling.orientation, "minMax")
        self.assertEqual(chart.series[0].tx.v, "Edited UV-Vis")
        self.assertEqual(
            chart.series[0].graphicalProperties.line.solidFill.srgbClr, "2468AC"
        )

    def test_gpc_excel_keeps_all_1501_unprocessed_ri_points(self):
        curve = load_gpc_csv(GPC_FOLDER / "GPC_RI_demo_07.csv")
        original_x = curve.retention_times_min
        original_y = curve.ri_signal_mv
        sheet, display = self._export(
            curve, GPC, AxisRange(0, 30, -1.0, 15.0), "gpc.xlsx"
        )
        chart = sheet._charts[0]
        self.assertEqual(display.x_values, original_x)
        self.assertEqual(display.y_values, original_y)
        self.assertEqual(sheet["A1"].value, "Edited GPC | RetentionTime_min")
        self.assertEqual(sheet["B1"].value, "Edited GPC | RI_mV")
        self.assertEqual(sheet.max_row, 1502)
        self.assertEqual(sheet["A1502"].value, original_x[-1])
        self.assertEqual(sheet["B1502"].value, original_y[-1])
        self.assertIn("'Data'!$A$2:$A$1502", chart.series[0].xVal.numRef.f)
        self.assertIn("'Data'!$B$2:$B$1502", chart.series[0].yVal.numRef.f)
        self.assertEqual(_axis_title(chart.x_axis), "Retention time (min)")
        self.assertEqual(_axis_title(chart.y_axis), "RI response (mV)")
        self.assertEqual(chart.x_axis.scaling.orientation, "minMax")
        self.assertEqual(chart.x_axis.numFmt.formatCode, "0.00")
        self.assertEqual(chart.series[0].tx.v, "Edited GPC")


if __name__ == "__main__":
    unittest.main()
