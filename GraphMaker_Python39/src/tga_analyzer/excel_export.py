from __future__ import annotations

from typing import Union

import os
import tempfile
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Font,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
)
from openpyxl.utils import get_column_letter

from .display_series import DisplaySeries, display_axis_titles, to_display_series
from .model import GPC, PARTICLE_SIZE, TGA, AxisRange, CurveData, PlotState
from .particle_size_processing import ParticleSizeProcessedData
from .processing import ProcessedCurveData

BLACK = "000000"
WHITE = "FFFFFF"
PLOT_BORDER_EMU = 19050
AXIS_LINE_EMU = 19050
DATA_HEADER_ROW = 1
DATA_START_ROW = DATA_HEADER_ROW + 1
DATA_COLUMN_WIDTH = 7
CHART_ANCHOR = "D5"
CHART_WIDTH_CM = 18
CHART_HEIGHT_CM = 12
AXIS_TITLE_FONT_PT = 14
TICK_LABEL_FONT_PT = 12
LEGEND_FONT_PT = 12
PLOT_LAYOUT_X = 0.09
PLOT_LAYOUT_Y = 0.02
PLOT_LAYOUT_W = 0.78
PLOT_LAYOUT_H = 0.87


def export_excel(
    curves: Sequence[Union[CurveData, ProcessedCurveData, ParticleSizeProcessedData, DisplaySeries]],
    plot_state: PlotState,
    output_path: Union[Path, str],
) -> Path:
    if not curves:
        raise ValueError("Excelへ出力する曲線がありません。")
    display_curves = tuple(to_display_series(curve) for curve in curves)
    if any(curve.measurement_type != plot_state.measurement_type for curve in display_curves):
        raise ValueError("異なる測定モードの曲線は同じExcelへ出力できません。")

    target = Path(output_path)
    if target.suffix.lower() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.freeze_panes = f"A{DATA_START_ROW}"
    data_sheet.sheet_view.showGridLines = False

    start_col = 1
    for curve in display_curves:
        headers, columns, formats = _curve_table(curve)
        for offset, header in enumerate(headers):
            data_sheet.cell(row=DATA_HEADER_ROW, column=start_col + offset, value=header)
        for point_index in range(curve.point_count):
            row = DATA_START_ROW + point_index
            for offset, values in enumerate(columns):
                data_sheet.cell(row=row, column=start_col + offset, value=values[point_index])
        max_row = curve.point_count + DATA_HEADER_ROW
        for offset, number_format in enumerate(formats):
            column = start_col + offset
            data_sheet.column_dimensions[get_column_letter(column)].width = DATA_COLUMN_WIDTH
            for row in range(DATA_START_ROW, max_row + 1):
                data_sheet.cell(row=row, column=column).number_format = number_format
        start_col += len(headers)

    chart = _build_chart(
        data_sheet, display_curves, plot_state.axis_range, plot_state.measurement_type
    )
    # Keep the native editable chart floating over the data cells.
    data_sheet.add_chart(chart, CHART_ANCHOR)
    mode = plot_state.measurement_type
    workbook.properties.title = f"{mode} Comparison"
    workbook.properties.subject = f"Editable {mode} overlay chart"
    workbook.properties.creator = "GraphMaker Python 3.9 Compatible"

    temp_handle = tempfile.NamedTemporaryFile(
        prefix=".thermal_export_", suffix=".xlsx", dir=target.parent, delete=False
    )
    temp_path = Path(temp_handle.name)
    temp_handle.close()
    try:
        workbook.save(temp_path)
        os.replace(temp_path, target)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return target


def _curve_table(
    curve: DisplaySeries,
) -> tuple[tuple[str, ...], tuple[tuple[float, ...], ...], tuple[str, ...]]:
    metadata = " | " + " | ".join(curve.header_metadata) if curve.header_metadata else ""
    columns = (
        (curve.x_data_header, curve.x_values, curve.x_number_format),
        (curve.y_data_header, curve.y_values, curve.y_number_format),
        *((column.header, column.values, column.number_format) for column in curve.extra_columns),
    )
    return (
        tuple(f"{curve.legend_name} | {header}{metadata}" for header, _values, _fmt in columns),
        tuple(values for _header, values, _fmt in columns),
        tuple(number_format for _header, _values, number_format in columns),
    )


def _build_chart(
    data_sheet,
    curves: Sequence[DisplaySeries],
    axis_range: AxisRange,
    measurement_type: str,
) -> ScatterChart:
    axis_range.validate()
    chart = ScatterChart()
    chart.height = CHART_HEIGHT_CM
    chart.width = CHART_WIDTH_CM
    chart.scatterStyle = "line"
    chart.varyColors = False
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=PLOT_LAYOUT_X,
            y=PLOT_LAYOUT_Y,
            w=PLOT_LAYOUT_W,
            h=PLOT_LAYOUT_H,
            xMode="factor",
            yMode="factor",
            wMode="factor",
            hMode="factor",
        )
    )
    chart.graphical_properties = GraphicalProperties(
        solidFill=WHITE,
        ln=LineProperties(noFill=True),
    )
    chart.plot_area.graphicalProperties = GraphicalProperties(
        solidFill=WHITE,
        ln=LineProperties(solidFill=BLACK, w=PLOT_BORDER_EMU),
    )
    x_title, y_title = display_axis_titles(curves, measurement_type)
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.scaling.min = axis_range.x_min
    chart.x_axis.scaling.max = axis_range.x_max
    chart.y_axis.scaling.min = axis_range.y_min
    chart.y_axis.scaling.max = axis_range.y_max
    reverse_x = curves[0].reverse_x
    if any(curve.reverse_x != reverse_x for curve in curves):
        raise ValueError("X軸方向が異なる系列は同じグラフへ出力できません。")
    chart.x_axis.scaling.orientation = "maxMin" if reverse_x else "minMax"
    logarithmic_x = curves[0].logarithmic_x
    if any(curve.logarithmic_x != logarithmic_x for curve in curves):
        raise ValueError("X軸形式が異なる系列を同じグラフへ出力できません。")
    if logarithmic_x:
        if axis_range.x_min <= 0 or axis_range.x_max <= 0:
            raise ValueError("対数X軸の最小値と最大値は0より大きい値にしてください。")
        chart.x_axis.scaling.logBase = 10
    chart.y_axis.scaling.orientation = "minMax"
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "max" if reverse_x else "min"
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorTickMark = "in"
    chart.y_axis.majorTickMark = "in"
    chart.x_axis.minorTickMark = "none"
    chart.y_axis.minorTickMark = "none"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.numFmt = (
        "0.00"
        if measurement_type == GPC
        else "0.######"
        if measurement_type == PARTICLE_SIZE
        else "0"
    )
    chart.y_axis.numFmt = "0" if measurement_type == TGA else "0.00"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=BLACK, w=AXIS_LINE_EMU)
    )
    chart.y_axis.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=BLACK, w=AXIS_LINE_EMU)
    )
    chart.legend.position = "r"
    chart.legend.overlay = False
    chart.display_blanks = "gap"
    _style_chart_text(chart)

    start_col = 1
    for curve in curves:
        max_row = curve.point_count + DATA_HEADER_ROW
        x_values = Reference(
            data_sheet, min_col=start_col, min_row=DATA_START_ROW, max_row=max_row
        )
        y_values = Reference(
            data_sheet,
            min_col=start_col + 1,
            min_row=DATA_START_ROW,
            max_row=max_row,
        )
        series = Series(y_values, x_values, title=curve.legend_name)
        series.marker.symbol = "none"
        series.graphicalProperties.line.solidFill = curve.color.lstrip("#")
        series.graphicalProperties.line.width = 19050
        series.smooth = False
        chart.series.append(series)
        start_col += len(_curve_table(curve)[0])
    return chart


def _character_properties(font_size_pt: int, *, bold: bool = False) -> CharacterProperties:
    return CharacterProperties(
        sz=font_size_pt * 100,
        b=bold,
        solidFill=BLACK,
        latin=Font(typeface="Arial"),
        ea=Font(typeface="Arial"),
    )


def _rich_text_properties(font_size_pt: int, *, bold: bool = False) -> RichText:
    properties = _character_properties(font_size_pt, bold=bold)
    return RichText(
        p=[
            Paragraph(
                pPr=ParagraphProperties(defRPr=properties),
                r=[RegularTextRun(rPr=properties, t="")],
                endParaRPr=properties,
            )
        ]
    )


def _style_title(title, font_size_pt: int, *, bold: bool = False) -> None:
    if title is None:
        return
    properties = _character_properties(font_size_pt, bold=bold)
    for paragraph in title.tx.rich.paragraphs:
        if paragraph.pPr is None:
            paragraph.pPr = ParagraphProperties()
        paragraph.pPr.defRPr = properties
        paragraph.endParaRPr = properties
        for run in paragraph.r:
            run.rPr = properties
    title.txPr = _rich_text_properties(font_size_pt, bold=bold)


def _style_chart_text(chart: ScatterChart) -> None:
    _style_title(chart.x_axis.title, AXIS_TITLE_FONT_PT, bold=True)
    _style_title(chart.y_axis.title, AXIS_TITLE_FONT_PT, bold=True)
    chart.x_axis.txPr = _rich_text_properties(TICK_LABEL_FONT_PT, bold=False)
    chart.y_axis.txPr = _rich_text_properties(TICK_LABEL_FONT_PT, bold=False)
    chart.legend.txPr = _rich_text_properties(LEGEND_FONT_PT, bold=False)
